from django.dispatch import dispatcher
from django.utils import timezone
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.urls import reverse_lazy
from django.views import generic
from openpyxl import Workbook

from .forms import *
from .models import *
from .models import Snippet
import account.forms
import account.views
import random
import string


def index(request):
    reviews = Review.objects.filter(published_date__lte=timezone.now()).order_by('published_date')
    return render(request, 'main/landing.html', {'reviews': reviews})


# ПОЛЬЗОВАТЕЛЬСКИЕ ФУНКЦИИ

class LoginView(account.views.LoginView):
    form_class = account.forms.LoginEmailForm
    

def sign_in(request):
    form = SignInForm
    return render(request, 'account/login.html', {'form': form})


class SignUp(generic.CreateView):
    form_class = SignUpForm
    success_url = reverse_lazy('login')
    template_name = 'account/signup.html'

    def generate_username(self, form):
        username = form.cleaned_data["email"]
        return username


def profile(request, pk):
    user = User.objects.get(pk=pk)
    users_list = User.objects.all()
    driver_lvl = ''
    cars = []
    l = []
    for g in request.user.groups.all():
        l.append(g.name)
    its_driver = 0
    if is_driver(user):
        its_driver = 1
        driver = DriverProfile.objects.get(user=user)
        cars = Car.objects.filter(owner=driver)
        if driver.driver_rating > 100:
            driver_lvl = 'Бронза'
        else:
            driver_lvl = 'Новичок'
    else:
        its_driver = 0
        driver = []
    profile = get_object_or_404(User, pk=pk)
    orders = Order.objects.filter(author=user)
    return render(request, 'user/profile.html', {
        'orders': orders,
        'users_list': users_list,
        'profile': profile,
        'user': user,
        'its_driver': its_driver,
        'driver': driver,
        'cars': cars,
        'driver_lvl': driver_lvl,
    })


@login_required
def view_user_profile(request, user_pk):
    user = User.objects.get(pk=user_pk)
    return render(request, 'profile.html')


def job_application(request):
    if request.method == 'POST':
        application_form = JobApplicationForm(request.POST)
        application_form.user = request.user
        if application_form.is_valid():
            application = application_form.save(commit=False)
            application.save()
            return redirect(index)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки.')
            return redirect(job_application)
    else:
        application_form = JobApplicationForm
    return render(request, 'user/job_application.html', {
        'application_form': application_form
    })


@login_required
@transaction.atomic
def update_profile(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return render(request, 'user/profile.html')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки.')
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)
    return render(request, 'user/profile_edit.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })


# ЗАКАЗЫ
def order(request):
    error = ''
    if request.method == 'POST':
        form = OrderForm(request.POST)
        user = request.user
        if form.is_valid():
            order = form.save(commit=False)
            order.sended_in = datetime.now()
            if user.is_authenticated:
                order.author = user
            size=8
            order.identificator=(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(size)))
            order.save()
            return ordered(request, pk=order.id)
        else:
            error = 'Форма была неверной'

    all_address = Address.objects.all()
    all_cityprice = CityPrice.objects.all()
    all_types = CargoType.objects.all()
    assistant_library = AssistantLibrary.objects.all()
    form = OrderForm()
    form.use_required_attribute = False

    data = {
        'form': form,
        'error': error,
        'all_address': all_address,
        'all_cityprice': all_cityprice,
        'all_types': all_types,
        'assistant_library': assistant_library,
    }
    return render(request, 'order/order.html', data)


def order_view(request, identificator):
    order_object = get_object_or_404(Order, identificator=identificator)
    return render(request, 'order/order_view.html', {'order': order_object})

def order_create(request):
    error = ''
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ordered')
        else:
            error = 'Форма была неверной'

    all_address = Address.objects.all()
    all_cityprice = CityPrice.objects.all()
    all_types = CargoType.objects.all()
    form = OrderForm()
    form.use_required_attribute = False

    data = {
        'form': form,
        'error': error,
        'all_address': all_address,
        'all_cityprice': all_cityprice,
        'all_types': all_types,
    }
    return render(request, 'order/order_create.html', data)


def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect(order_detail, pk=order.pk)

    else:
        form = OrderEditForm(instance=order)

    return render(request, 'order/order_edit.html', {'form': form})


def order_delete(request, pk):
    try:
        order = get_object_or_404(Order, pk=pk)
        if order.driver and order.status<3:
            driver = DriverProfile.objects.get(driver=order.driver)
            driver.current_orders-=1
            driver.save()
        order.delete()
        redirect_url = reverse(order_table)
        return redirect(redirect_url)

    except ValueError:
        return render(request, 'order/order_table.html')


def ordered(request, pk):
    order_object = get_object_or_404(Order, pk=pk)
    return render(request, 'order/ordered.html', {'order': order_object})


def report(request):
    report = ReportForm(request.POST)
    if request.method == 'POST':
        if report.is_valid():
            report.save()
            return redirect('index')
        else:
            error = 'Форма была неверной'
    return render(request, 'help/report.html', {'report': report})


def myorders(request):
    orders = Order.objects.all()
    return render(request, 'driver/driver_profile.html', {'orders': orders})


def order_enable(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if not order.driver == None:
        order.status = 2
        order.save()
    else:
        messages.info(request, 'К данному заказу не присвоен водитель, добавьте водителя к заказу для продолжения.')

    return order_detail(request, pk)


def order_disable(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order.status = 1
    order.save()
    return order_detail(request, pk)


def order_complete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.driver:
        order.status = 3
        driver = DriverProfile.objects.get(driver=order.driver)
        driver.current_orders-=1
        order.save()
        driver.save()
    else:
        messages.info(request, 'К данному заказу не присвоен водитель, добавьте водителя к заказу для продолжения.')
    return redirect(order_detail, pk=order.pk)


def order_drivers(request, pk):
    order = get_object_or_404(Order, pk=pk)
    orders = Order.objects.all()
    drivers=DriverProfile.objects.all().order_by('-driver_rating')
    users=Profile.objects.all()
    for driver in drivers:
        driver_orders=0
        for order in orders:
            if order.driver == driver:
                driver_orders+=1
        if not driver_orders == driver.current_orders:
            driver.current_orders=driver_orders
            driver.save()
    if DispatcherProfile.objects.filter(user=request.user):
        if order.dispatcher.user == request.user:
            is_dispatcher=True
        else:
            is_dispatcher=False  
    else:
        is_dispatcher=False
    return render(request, 'order/order_drivers.html', {'order': order, 'drivers': drivers, 'is_dispatcher': is_dispatcher})


def order_driver_select(request, pk, driver_pk):
    order = get_object_or_404(Order, pk=pk)
    driver=DriverProfile.objects.get(pk=driver_pk)
    order.driver = driver
    driver.current_orders+=1
    driver.save()
    order.save()
    return order_detail(request, pk)


def order_dispatchers(request, pk):
    order = get_object_or_404(Order, pk=pk)
    dispatchers=DispatcherProfile.objects.all().order_by('-dispatcher_rating')
    return render(request, 'order/order_dispatchers.html', {'order': order, 'dispatchers': dispatchers})


def order_dispatcher_select(request, pk, dispatcher_pk):
    order = get_object_or_404(Order, pk=pk)
    order.dispatcher = DispatcherProfile.objects.get(pk=dispatcher_pk)
    order.status=1
    order.save()
    return order_detail(request, pk)


def order_for_dispatcher(request):
    dispatchers=DispatcherProfile.objects.all()
    dispatcher = {}
    for dispatch in dispatchers:
        if dispatch.user == request.user:
            dispatcher=dispatch
    if dispatcher:
        dispatcher_orders=Order.objects.filter(dispatcher=dispatcher)
        for order in dispatcher_orders:
            if order.status<2:
                messages.info(request, 'У вас есть не обработанные заказы')
                pk=order.id
                return order_detail(request, pk)
        if(Order.objects.filter(dispatcher=None).order_by('-sended_in') and Order.objects.filter(status=0).order_by('-sended_in')[0]):
            order=Order.objects.filter(dispatcher=None).order_by('-sended_in')
            order=Order.objects.filter(status=0).order_by('-sended_in')[0]
            order.dispatcher=dispatcher
            order.status=1
            order.save()
            pk=order.id
            return order_detail(request, pk)
        else:
            messages.info(request, 'Нет текущих заказов для вас')
            return order_table(request)

    else:
        messages.info(request, 'Вы не диспетчер')
        return order_table(request)



def order_table(request):
    user = request.user
    all_types = CargoType.objects.all()

    if user.is_superuser:
        orders = Order.objects.all().order_by('start_time')
        drivers = DriverProfile.objects.all()
        return render(request, 'order/order_table.html', {
            'orders': orders, 'all_types': all_types, 'drivers': drivers})

    elif DispatcherProfile.objects.filter(user=user):
        drivers = DriverProfile.objects.all()
        dispatcher=DispatcherProfile.objects.get(user=request.user)
        orders = Order.objects.filter(dispatcher=dispatcher).order_by('-sended_in')
        return render(request, 'order/order_table.html',
                      {'orders': orders, 'all_types': all_types, 'drivers': drivers})

    elif DriverProfile.objects.filter(user=user):
        driver = DriverProfile.objects.get(user=user)
        orders = Order.objects.filter(driver=driver)
        return render(request, 'order/order_table.html',
                      {'orders': orders, 'all_types': all_types})

    else:
        orders = Order.objects.filter(author=user)
        drivers = DriverProfile.objects.all()
        return render(request, 'order/order_table.html', {
            'orders': orders, 'all_types': all_types, 'drivers': drivers})


def driver_cars(request, pk):
    driver = DriverProfile.objects.get(pk=pk)
    cars = Car.objects.filter(owner=driver)
    return render(request, 'driver/driver_cars.html', {'driver': driver, 'cars': cars})


def order_detail(request, pk):
    permission = False
    order_object = get_object_or_404(Order, pk=pk)
    if request.user.is_superuser:
        permission = True
    if order_object.driver:
        if order_object.driver.user == request.user:
            permission = True
    if order_object.author:
        if order_object.author == request.user:
            permission = True
    if request.user.is_superuser:
        permission = True
    if DispatcherProfile.objects.filter(user=request.user):
        permission = True
        is_dispatcher=True
    else:
        is_dispatcher=False
    return render(request, 'order/order_detail.html', {'order': order_object, 'permission': permission, 'is_dispatcher': is_dispatcher})


def set_dispatcher():
    pass


def is_driver(user):
    return user.groups.filter(name='Водитель').exists()


def is_dispatcher(user):
    return user.groups.filter(name='Диспетчер').exists()


# ОТЗЫВЫ
@login_required
def review_create(request):
    if request.method == 'POST':
        review_form = ReviewForm(request.POST)
        if review_form.is_valid():
            review = review_form.save(commit=False)
            review.author = request.user
            review.published_date = timezone.now()
            review.save()
            return redirect(index)
        else:
            return redirect(review_create)
    else:
        review_form = ReviewForm
    return render(request, 'main/review.html', {
        'review_form': review_form
    })


def index(request):
    if request.method == 'POST':
        review_form = ReviewForm(request.POST)
        telorder = TelOrderForm(request.POST)
        if telorder.is_valid():
            telorder.save()
            return redirect(ordered)
        else:
            if review_form.is_valid():
                review = review_form.save(commit=False)
                review.author = request.user
                review.published_date = timezone.now()
                review.save()
                return redirect(index)
            else:
                return redirect(index)
    else:
        review_form = ReviewForm
        telorder = TelOrderForm

    reviews = Review.objects.filter(published_date__lte=timezone.now()).order_by('published_date')
    return render(request, 'main/landing.html', {'reviews': reviews, 'review_form': review_form, 'telorder': telorder})


def licenses(request):
    return render(request, 'licenses/licenses.html', {})


def orders_export(request):
    order_queryset = Order.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-orders.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Заказы'

    # Define the titles for columns
    columns = [
        'ID',
        'Отправитель',
        'Номер телефона',
        'Маршрут',
        'Грузчики, чел.',
        'Грузчики, час.',
        'Вид груза',
        'Сумма',
        'Статус',

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all orders
    for order in order_queryset:
        row_num += 1
        order.start_time = str(order.start_time)
        # Define the data for each cell in the row 
        row = [
            order.id,
            order.author.username if order.author else None,
            order.user_tel_nomer,
            order.full_road,
            order.loader_count,
            order.loader_time_count,
            order.cargo_type,
            order.order_price,
            order.status,
        ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
  
  
def about(request):
    return render(request, 'main/about.html', {})


def yandex_66b5cc356c187df1(request):
    return render(request, 'main/yandex_66b5cc356c187df1.html', {})


def snippet_detail(request, slug):
    snippet = get_object_or_404(Snippet, slug=slug)
    return HttpResponse(f'This should be the detail view for the slug of {slug}')


def company(request, pk):
    company = Company.objects.get(pk=pk)
    company_orders = CompanyOrder.objects.filter(
        company=company
    )
    return render(request, 'company/company_profile.html', {
        'company_orders': company_orders,
        })


def company_create(request):
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.contact_user=request.user
            form.save()
            return render(request, 'company/company_profile.html', {'pk': form.pk})
    form=CompanyForm()
    return render(request, 'company/company_create.html', {'form': form})
    


def company_edit(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            return render(request, 'company/company_profile.html', {'pk': company.pk})


def company_order_driver_select(request, pk, driver_pk):
    order = get_object_or_404(CompanyOrder, pk=pk)
    order.driver = DriverProfile.objects.get(pk=driver_pk)
    order.save()
    return order_detail(request, pk)


def company_order_detail(request, pk):
    permission = False
    order_object = get_object_or_404(CompanyOrder, pk=pk)
    if request.user.is_superuser:
        permission = True
    if order_object.driver:
        if order_object.driver.user == request.user:
            permission = True
    if order_object.company.contact_user == request.user:
        permission = True
    if DispatcherProfile.objects.filter(user=request.user):
        permission = True
    return render(request, 'company_order/company_order_detail.html', {'order': order_object, 'permission': permission})


def company_order_create(request):
    if request.method == 'POST':
        form = CompanyOrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ordered')



def company_order_edit(request, pk):
    company_order = get_object_or_404(CompanyOrder, pk=pk)
    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=company_order)
        if form.is_valid():
            form.save()
            return redirect(order_detail, pk=company_order.pk)


def company_order_delete(request, pk):
    try:
        company_order = get_object_or_404(CompanyOrder, pk=pk)
        company_order.delete()
        return redirect(reverse(order_table))

    except ValueError:
        return render(request, 'company/company_profile.html')


def company_orders(request):
    orders = CompanyOrder.objects.filter(
        company__contact_user=request.user
    )
    return render(request, 'company/company_profile.html', {'orders': orders})
