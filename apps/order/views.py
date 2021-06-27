from apps import order
from django.dispatch import dispatcher
from django.utils import timezone
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.urls import reverse_lazy
from django.views import generic
from openpyxl import Workbook

from order.models import *
from order.forms import *
from order.views  import *
from users.forms import *
from core.models import *
from .forms import *
from .models import *
import random
import string

# Create your views here.
def order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        user = request.user
        if form.is_valid():
            order = form.save(commit=False)
            order.sended_in = datetime.now()
            if user.is_authenticated:
                order.author = user
            size=8
            order.identificator=(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(size)))
            order.save()
            return ordered(request, pk=order.id)
        else:
            messages.info(request, 'Не все поля формы заполнены')

    all_address = Address.objects.all()
    all_cityprice = CityPrice.objects.all()
    all_types = CargoType.objects.all()
    assistant_library = AssistantLibrary.objects.all()
    form = OrderForm()
    form.use_required_attribute = False
    data = {
        'form': form,
        'all_address': all_address,
        'all_cityprice': all_cityprice,
        'all_types': all_types,
        'assistant_library': assistant_library,
    }
    return render(request, 'order/order.html', data)


def order_view(request, identificator):
    order_object = get_object_or_404(Order, identificator=identificator)
    return render(request, 'order/order_view.html', {'order': order_object})

def order_create(request):
    error = ''
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ordered')
        else:
            error = 'Форма была неверной'

    all_address = Address.objects.all()
    all_cityprice = CityPrice.objects.all()
    all_types = CargoType.objects.all()
    form = OrderForm()
    form.use_required_attribute = False

    data = {
        'form': form,
        'error': error,
        'all_address': all_address,
        'all_cityprice': all_cityprice,
        'all_types': all_types,
    }
    return render(request, 'order/order_create.html', data)


def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect(order_detail, pk=order.pk)

    else:
        form = OrderEditForm(instance=order)

    return render(request, 'order/order_edit.html', {'form': form})


def order_delete(request, pk):
    try:
        order = get_object_or_404(Order, pk=pk)
        order.delete()
        redirect_url = reverse(order_table)
        return redirect(redirect_url)

    except ValueError:
        return render(request, 'order/order_table.html')


def ordered(request, pk):
    order_object = get_object_or_404(Order, pk=pk)
    return render(request, 'order/ordered.html', {'order': order_object})


def tel_ordered(request):
    return render(request, 'order/ordered_telephone.html')


def myorders(request):
    orders = Order.objects.all()
    return render(request, 'driver/driver_profile.html', {'orders': orders})


def order_enable(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if not order.driver == None:
        order.status = 2
        order.save()
    else:
        messages.info(request, 'К данному заказу не присвоен водитель, добавьте водителя к заказу для продолжения.')

    return order_detail(request, pk)


def order_disable(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order.status = 1
    order.save()
    return order_detail(request, pk)


def order_cancel(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order.status = -1
    order.save()
    return order_detail(request, pk)


def order_refresh(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order.status = 0
    order.save()
    return order_detail(request, pk)


def order_complete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.driver:
        order.status = 3
        order.save()
    else:
        messages.info(request, 'К данному заказу не присвоен водитель, добавьте водителя к заказу для продолжения.')
    return redirect(order_detail, pk=order.pk)


def order_drivers(request, pk):
    order = get_object_or_404(Order, pk=pk)
    orders = Order.objects.all()
    drivers=DriverProfile.objects.all().order_by('-driver_rating')
    users=Profile.objects.all()
    for driver in drivers:
        driver_orders=0
        for order in orders:
            if order.driver == driver and not order.status == 3:
                driver_orders+=1
        if not driver_orders == driver.current_orders:
            driver.current_orders=driver_orders
            driver.save()
    if DispatcherProfile.objects.filter(user=request.user):
        if order.dispatcher.user == request.user:
            is_dispatcher=True
        else:
            is_dispatcher=False  
    else:
        is_dispatcher=False
    return render(request, 'order/order_drivers.html', {'order': order, 'drivers': drivers, 'is_dispatcher': is_dispatcher})


def order_driver_select(request, pk, driver_pk):
    order = get_object_or_404(Order, pk=pk)
    driver=DriverProfile.objects.get(pk=driver_pk)
    order.driver = driver
    driver.current_orders+=1
    driver.save()
    order.save()
    return order_detail(request, pk)


def order_dispatchers(request, pk):
    order = get_object_or_404(Order, pk=pk)
    dispatchers=DispatcherProfile.objects.all().order_by('-dispatcher_rating')
    return render(request, 'order/order_dispatchers.html', {'order': order, 'dispatchers': dispatchers})


def order_dispatcher_select(request, pk, dispatcher_pk):
    order = get_object_or_404(Order, pk=pk)
    order.dispatcher = DispatcherProfile.objects.get(pk=dispatcher_pk)
    order.status=1
    order.save()
    return order_detail(request, pk)


def order_for_dispatcher(request):
    if DispatcherProfile.objects.filter(user=request.user):
        dispatcher_orders=Order.objects.filter(dispatcher=dispatcher)
        for order in dispatcher_orders:
            if order.status<2:
                messages.info(request, 'У вас есть не обработанные заказы')
                pk=order.id
                return order_detail(request, pk)
        if(Order.objects.filter(dispatcher=None).order_by('-sended_in') and Order.objects.filter(status=0).order_by('-sended_in')[0]):
            order=Order.objects.filter(dispatcher=None).order_by('-sended_in')
            order=Order.objects.filter(status=0).order_by('-sended_in')[0]
            order.dispatcher=dispatcher
            order.status=1
            order.save()
            pk=order.id
            return order_detail(request, pk)
        else:
            messages.info(request, 'Нет текущих заказов для вас')
            return order_table(request)

    else:
        messages.info(request, 'Вы не диспетчер')
        return order_table(request)



def order_table(request):
    user = request.user
    all_types = CargoType.objects.all()
    if user.is_superuser:
        orders = Order.objects.all().order_by('start_time')
        drivers = DriverProfile.objects.all()
        return render(request, 'order/order_table.html', {
            'orders': orders, 'all_types': all_types, 'drivers': drivers})

    elif DispatcherProfile.objects.filter(user=user):
        drivers = DriverProfile.objects.all()
        dispatcher=DispatcherProfile.objects.get(user=request.user)
        orders = Order.objects.filter(dispatcher=dispatcher).order_by('-sended_in')
        return render(request, 'order/order_table.html',
                      {'orders': orders, 'all_types': all_types, 'drivers': drivers})

    elif DriverProfile.objects.filter(user=user):
        driver = DriverProfile.objects.get(user=user)
        orders = Order.objects.filter(driver=driver)
        return render(request, 'order/order_table.html',
                      {'orders': orders, 'all_types': all_types})

    else:
        orders = Order.objects.filter(author=user)
        drivers = DriverProfile.objects.all()
        return render(request, 'order/order_table.html', {
            'orders': orders, 'all_types': all_types, 'drivers': drivers})


def driver_cars(request, pk):
    driver = DriverProfile.objects.get(pk=pk)
    cars = Car.objects.filter(owner=driver)
    return render(request, 'driver/driver_cars.html', {'driver': driver, 'cars': cars})


def order_detail(request, pk):
    permission = False
    order_object = get_object_or_404(Order, pk=pk)
    if request.user.is_superuser:
        permission = True
    if order_object.driver:
        if order_object.driver.user == request.user:
            permission = True
    if order_object.author:
        if order_object.author == request.user:
            permission = True
    if request.user.is_superuser:
        permission = True
    if DispatcherProfile.objects.filter(user=request.user):
        permission = True
        is_dispatcher=True
    else:
        is_dispatcher=False
    return render(request, 'order/order_detail.html', {'order': order_object, 'permission': permission, 'is_dispatcher': is_dispatcher})


def orders_export(request):
    order_queryset = Order.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-orders.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Заказы'

    # Define the titles for columns
    columns = [
        'ID',
        'Отправитель',
        'Номер телефона',
        'Маршрут',
        'Грузчики, чел.',
        'Грузчики, час.',
        'Вид груза',
        'Сумма',
        'Статус',

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all orders
    for order in order_queryset:
        row_num += 1
        order.start_time = str(order.start_time)
        # Define the data for each cell in the row 
        row = [
            order.id,
            order.author.username if order.author else None,
            order.user_tel_nomer,
            order.full_road,
            order.loader_count,
            order.loader_time_count,
            order.cargo_type,
            order.order_price,
            order.status,
        ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response